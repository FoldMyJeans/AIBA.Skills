"""Run the ba-user-stories review pass over a story workbook.

    python scripts/validate_workbook.py <workbook.xlsx|export.csv> [--kb <knowledgebase dir>]
    python scripts/validate_workbook.py <matrix.xlsx> --matrix
    python scripts/validate_workbook.py --selfcheck

The review pass in skills/ba-user-stories/SKILL.md is the specification; this
only automates the mechanical half of it. Judgment stays with the BA: whether a
story is right, whether an assumption is safe, whether the criteria describe the
business. Findings marked BLOCKING corrupt the record or break a gate, and the
exit code is 1 when any are present. Heuristics never block.

Reads the security tier and the formality from 01_Requirements_Source.md when the
knowledgebase is given, so the tier 2 NFR check knows which bar applies and
Contract can bar what Firm allows.
"""

import argparse
import contextlib
import csv
import datetime as dt
import io
import re
import sys
import tempfile
from collections import defaultdict
from functools import cache
from pathlib import Path

import openpyxl

FIELDS = [
    "Story ID", "Date Created", "Type", "Epic", "Phase", "User Role",
    "User Story", "Description / Notes", "Acceptance Criteria",
    "Priority (MoSCoW)", "Data Source", "Dependencies", "Assumptions",
    "Open Questions", "Status", "Approved By",
]

# Epic, User Role and Data Source are closed too, but their values live in 01
# and are not parsed here, see the note in ba-user-stories.
# Lists, not sets: the declared order is the lifecycle, and the templates build
# their dropdowns straight from these.
CLOSED = {
    "Type": ["Story", "NFR", "Constraint"],
    "Phase": ["Phase 1", "Phase 2", "Backlog", "Out of Scope"],
    "Priority (MoSCoW)": ["Must", "Should", "Could", "Won't"],
    "Status": ["Draft", "In Review", "Approved", "In Development", "In QA",
               "Done", "Deferred", "Deprecated"],
}

VAGUE_LIST = (Path(__file__).resolve().parent.parent
              / "skills" / "ba-user-stories" / "references" / "vague-words.md")

# The compliance matrix (ba-bid-response). Client requirement IDs are theirs.
MATRIX_FIELDS = ["Requirement ID", "Requirement", "Response", "How", "Assumption", "Reference"]
MATRIX_RESPONSES = ["Comply", "Partial", "Alternative", "Cannot"]


@cache
def vague_words():
    """The shared list, read from its one home. Empty when it cannot be read."""
    if not VAGUE_LIST.exists():
        return ()
    body = VAGUE_LIST.read_text(encoding="utf-8", errors="replace")
    blocks = re.findall(r"```[a-z]*\n(.*?)```", body, re.S)
    if not blocks:
        return ()
    return tuple(w.strip() for w in blocks[-1].splitlines()
                 if w.strip() and not w.lstrip().startswith("#"))

# Only the two categories the NFR catalog calls non negotiable on tier 2.
NON_NEGOTIABLE = {
    "security": ["security", "auth", "mfa", "sso", "encrypt", "tls", "password",
                 "lockout", "session", "permission"],
    "integration resilience": ["timeout", "retry", "unreachable", "unavailable",
                               "integration", "circuit breaker"],
}

FAILURE_HINTS = ["error", "fail", "invalid", "missing", "unavailable", "denied",
                 "timeout", "retry", "empty", "blank", "duplicate", "expired",
                 "revoked", "rejected", "unreachable", "offline", "conflict"]

# The grammar 02_Conventions.md declares: 2026-05-06, 6 May 2026, May 6, 2026.
# Slash dates are deliberately absent, because 06/05/2026 is two different days
# depending on who typed it, and this record gets read in a dispute. Anchored to
# real month names, or "revisit at sprint 6 2026" becomes a date that will not
# parse, which is the blocking false positive this file exists to avoid.
MONTH = r"(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*"
DATE_TOKEN = (r"\d{4}-\d{2}-\d{2}"
              rf"|\d{{1,2}}\s+{MONTH}\s+\d{{4}}"
              rf"|{MONTH}\s+\d{{1,2}},?\s+\d{{4}}")
DATE_FORMATS = ["%Y-%m-%d", "%d %b %Y", "%d %B %Y", "%b %d %Y", "%B %d %Y"]
DATE_HELP = "dates as 2026-05-06, 6 May 2026, or May 6, 2026"

# An annotation opens by pointing at the record it supersedes; an approver opens
# with a role. Anchored, because an unanchored match would let "Finance Lead
# pending, see 04" launder itself past the missing-date check.
ANNOTATION = re.compile(r"^\s*(supersed|see 04)", re.I)

# The marker ba-project-intake prescribes, punctuation-tolerant because it is
# stated in running prose. Matching the bare word would quarantine a story whose
# Assumptions mention unaudited management accounts.
QUARANTINE_MARKER = re.compile(r"unaudited\W+treat as unverified", re.I)

STORY_ID = r"[A-Za-z]{2,}-\d+"


def norm(text):
    """Excel autocorrects a typed apostrophe; Won't and Won<curly>t are one value."""
    return text.replace("’", "'").strip()


def parse_date(text):
    """First date-looking token in text, or None."""
    for raw in re.findall(DATE_TOKEN, text or ""):
        cleaned = re.sub(r"\s+", " ", raw).strip().replace(",", "")
        for fmt in DATE_FORMATS:
            try:
                return dt.datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def read_dial(kb_dir, pattern):
    """One dial's value from the Dials block of 01, or None.

    Scoped to the block: a Constraints paragraph mentioning a contract being
    negotiated must not set the formality of a project that left it unset.
    """
    source = Path(kb_dir) / "01_Requirements_Source.md"
    if not source.exists():
        return None
    text = source.read_text(encoding="utf-8", errors="replace").lower()
    block = re.search(r"^##\s*dials\b(.*?)(?=^##\s|\Z)", text, re.S | re.M)
    match = re.search(pattern, block.group(1) if block else text)
    return match.group(1) if match else None


def read_tier(kb_dir):
    """Security tier from the Dials block, or None."""
    value = read_dial(kb_dir, r"security tier\s*[:\-–—]?\s*\**\s*tier\s*(2\s*plus|2\s*\+|2|1)")
    if value is None:
        return None
    value = value.replace(" ", "")
    return "2+" if value in ("2plus", "2+") else value


def read_formality(kb_dir):
    """Formality from the Dials block, or None. Contract changes what is reported."""
    return read_dial(kb_dir, r"formality\s*[:\-–—]?\s*\**\s*(lean|firm|contract)")


def load_rows(path, key="Story ID"):
    """(rows as dicts keyed by header, header list, merged-cell count).

    A .csv is read as an export: a project whose authoritative store is the
    tracker still owes the same review pass, and the export is how it runs.
    """
    if str(path).lower().endswith(".csv"):
        # Excel's own CSV export is cp1252 on a Western box. Replacing the
        # undecodable byte turned a legal Won't into a blocking closed set
        # finding, so decode it properly instead: utf-8 first, cp1252 second.
        try:
            text = Path(path).read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            text = Path(path).read_text(encoding="cp1252", errors="replace")
        grid, merged = list(csv.reader(io.StringIO(text))), 0
    else:
        book = openpyxl.load_workbook(path, data_only=True)
        sheet = next((s for s in book
                      if any(str(c.value).strip() == key for c in s[1])), book.active)
        grid, merged = list(sheet.iter_rows(values_only=True)), len(sheet.merged_cells.ranges)
    if not grid:
        return [], [], 0
    headers = [str(h).strip() if h is not None else "" for h in grid[0]]
    rows = []
    for values in grid[1:]:
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        rows.append({h: ("" if v is None else str(v).strip())
                     for h, v in zip(headers, values) if h})
    return rows, headers, merged


def check(rows, headers, tier=None, merged=0, today=None, formality=None):
    """Return a list of (blocking, cause, story_id) findings."""
    today = today or dt.date.today()
    vague = vague_words()
    found = []

    def add(blocking, cause, story_id=""):
        found.append((blocking, cause, story_id))

    missing = [f for f in FIELDS if f not in headers]
    if missing:
        add(True, f"workbook is missing required columns: {', '.join(missing)}")
    named = [h for h in headers if h]
    if len(named) != len(set(named)):
        add(True, "workbook has duplicate column headers, which breaks parsing")
    if merged:
        add(True, f"{merged} merged cell range(s) in the story sheet, and one row is one story")

    debt_keys = set()
    ids = {r.get("Story ID", "").upper() for r in rows}
    quarantined = {r.get("Story ID", "").upper() for r in rows
                   if QUARANTINE_MARKER.search(r.get("Assumptions", ""))}
    seen = set()

    for row in rows:
        sid = row.get("Story ID", "") or "(no Story ID)"
        row_text = " ".join(row.get(f, "") for f in FIELDS if f in row)

        if sid in seen and row.get("Story ID", ""):
            add(True, "duplicate Story ID", sid)
        seen.add(sid)

        for field in FIELDS:
            # Approved By is the one field where blank is an answer: it means
            # not approved. Everywhere else a blank makes the reader guess.
            if field == "Approved By":
                continue
            if field in headers and not row.get(field, ""):
                add(False, f"blank {field}, write None or N/A, never nothing", sid)

        for field, allowed in CLOSED.items():
            value = norm(row.get(field, ""))
            if value and value not in allowed:
                add(True, f"{field} is outside its closed set: {value!r}", sid)

        for dep in re.findall(STORY_ID, row.get("Dependencies", "")):
            if dep.upper() not in ids:
                add(True, f"dependency {dep} does not exist in this workbook", sid)
            elif dep.upper() in quarantined:
                add(True, f"dependency {dep} points into a quarantined ticket, which "
                          f"must be audited first", sid)

        status = row.get("Status", "")
        approved_by = row.get("Approved By", "")
        at_risk = "proceeding at risk" in approved_by.lower()

        # Not blocking: the record itself is the honest artifact ba-change asks
        # for, and a shipped fix is allowed to carry one even here. What must not
        # happen, meaning a ticket or a signature, is not visible from a workbook.
        if at_risk and formality == "contract":
            add(False, "proceed at risk record on a Contract project, so it goes to the "
                       "spec's assumptions section or blocks signature, and never "
                       "creates a ticket (ba-change)", sid)
        if at_risk and status == "Approved":
            add(True, "proceed at risk record on an Approved row, so Status stays In Review", sid)
        elif status == "Approved":
            # An annotation with no date is a pointer to a record. An approval
            # always carries its date, so a dated segment is never an annotation
            # even when it explains itself ("covers work already built").
            segments = [s for s in approved_by.split(";") if s.strip()
                        and not (ANNOTATION.search(s) and parse_date(s) is None)]
            if not segments or any(parse_date(s) is None for s in segments):
                add(True, f"Status is Approved but an approver carries no date, and joint "
                          f"approvals need every role and its own date ({DATE_HELP})", sid)
            elif not any(re.search(r"[A-Za-z]", re.sub(DATE_TOKEN, "", s)) for s in segments):
                add(True, "Status is Approved but Approved By names no role", sid)

        # The date sits immediately after the word. Anything looser turned "the
        # owner asked us to revisit this wording" into a blocking finding on an
        # ordinary approved story, and a keyword match must never do that.
        stamp = re.search(rf"revisit\W{{1,3}}({DATE_TOKEN})", row_text, re.I)
        # The four instances ba-change names. A plain hold is not one of them: a
        # story parked by a directive is authorized, not owed. A deferred NFR is,
        # because that is the accessibility target deferred on an MVP clock.
        is_debt = (at_risk or sid.upper() in quarantined
                   or ("revisit" in approved_by.lower()
                       and re.search(r"\bCH-\d+\b", approved_by))
                   or (status == "Deferred" and row.get("Type", "") == "NFR"))

        if stamp:
            # One quarantine set is one record, not one per row.
            if is_debt:
                handle = re.search(r"\bCH-\d+\b", row_text)
                debt_keys.add("quarantine" if sid.upper() in quarantined
                              else handle.group(0) if handle else sid)
            revisit = parse_date(stamp.group(1))
            # Date shaped is not the same as a real date: 30 Feb 2026 matches the
            # pattern and parses to nothing, and a deferral nothing can expire is
            # the failure this whole check exists to prevent.
            if revisit is None:
                add(True, f"revisit date {stamp.group(1)!r} is not a real date ({DATE_HELP})", sid)
            elif revisit < today:
                add(True, f"debt record is past its revisit date ({revisit})", sid)
        elif is_debt:
            add(True, "debt record here carries no revisit date, so nothing brings "
                      "it back", sid)
        elif status == "Deferred" and row.get("Phase", "") != "Out of Scope":
            add(True, "Deferred with no revisit date, and a hold nothing brings back is "
                      "a story quietly dying", sid)

        haystack = " ".join(row.get(f, "") for f in
                            ("User Story", "Acceptance Criteria")).lower()
        for word in vague:
            if not re.search(rf"\b{re.escape(word)}\b", haystack):
                continue
            # The remedy the review pass asks for: a number, or an [OPEN].
            if "[open]" in haystack or re.search(rf"\b{re.escape(word)}\b\W+\S*\d", haystack):
                continue
            add(False, f"vague word {word!r}, so make it a number, a named control, "
                       f"an eval threshold, or an [OPEN]", sid)

        if (row.get("Priority (MoSCoW)", "") == "Must"
                and row.get("Phase", "") == "Phase 1"
                and row.get("Type", "") == "Story"):
            criteria = row.get("Acceptance Criteria", "").lower()
            if criteria and not any(re.search(rf"\b{h}", criteria) for h in FAILURE_HINTS):
                add(False, "Phase 1 Must with no failure path in its acceptance "
                           "criteria (heuristic, check by eye)", sid)

    # Each one can be individually defensible; the pile is the finding.
    if len(debt_keys) >= 3:
        for row in rows:
            if "revisit" in " ".join(row.get(f, "") for f in FIELDS if f in row).lower():
                add(False, f"{len(debt_keys)} open debt records on one project, so the work is "
                           f"outrunning the authority for it. Escalate the pattern, not the "
                           f"items (heuristic, check by eye)", row.get("Story ID", ""))

    if tier in ("2", "2+"):
        nfr_text = " ".join(r.get("User Story", "") + " " + r.get("Acceptance Criteria", "")
                            for r in rows if r.get("Type", "") == "NFR").lower()
        for category, words in NON_NEGOTIABLE.items():
            if not any(w in nfr_text for w in words):
                add(False, f"tier {tier} project has no {category} NFR row, and that "
                           f"category is not negotiable (heuristic, check by eye)")

    return found


def check_matrix(rows, headers, merged=0):
    """The compliance matrix rules from ba-bid-response."""
    vague = vague_words()
    found = []

    missing = [f for f in MATRIX_FIELDS if f not in headers]
    if missing:
        found.append((True, f"matrix is missing required columns: {', '.join(missing)}", ""))
    if merged:
        found.append((True, f"{merged} merged cell range(s), and the client's numbering is law, "
                            f"do not renumber and do not merge rows", ""))

    seen = set()
    for row in rows:
        rid = row.get("Requirement ID", "") or "(no Requirement ID)"
        response = norm(row.get("Response", ""))
        needs_assumption = response in ("Partial", "Alternative")

        if rid in seen and row.get("Requirement ID", ""):
            found.append((True, "duplicate Requirement ID, and the evaluator reads answers "
                                "against their own document", rid))
        seen.add(rid)

        if response and response not in MATRIX_RESPONSES:
            found.append((True, f"Response is outside its closed set: {response!r}, and "
                                f"evaluators filter this column", rid))
        elif not response:
            found.append((True, "blank Response, and every requirement gets one of "
                                "Comply, Partial, Alternative, Cannot", rid))
        for field in ("Requirement ID", "Requirement", "How", "Assumption", "Reference"):
            # The Partial rule below says it better; do not report the cell twice.
            if field == "Assumption" and needs_assumption:
                continue
            if field in headers and not row.get(field, ""):
                found.append((False, f"blank {field}, write None, never nothing", rid))
        if needs_assumption and \
                row.get("Assumption", "").strip().lower() in ("", "none", "n/a"):
            found.append((True, f"{response} with no assumption, and unstated assumptions "
                                f"become unpaid scope after award", rid))
        text = row.get("Requirement", "").lower()
        for word in vague:
            if re.search(rf"\b{re.escape(word)}\b", text):
                found.append((False, f"vague word {word!r} in the client's requirement, so "
                                     f"price it with an assumption naming what was assumed", rid))

    return found


def report(found):
    """Print findings, collapsing rows that fail for the identical cause."""
    if not found:
        print("Review pass clean. The judgment half is still yours: read the "
              "criteria, not just the cells.")
        return 0

    grouped = defaultdict(list)
    for blocking, cause, sid in found:
        grouped[(blocking, cause)].append(sid)

    blocking_count = 0
    for (blocking, cause), sids in sorted(grouped.items(), key=lambda kv: not kv[0][0]):
        blocking_count += len(sids) if blocking else 0
        named = [s for s in sids if s]
        if len(named) > 6:
            where = f"{len(named)} rows: {', '.join(named[:6])} and {len(named) - 6} more"
        else:
            where = ", ".join(named)
        print(("[BLOCKING] " if blocking else "") + cause + (f" ({where})" if where else ""))

    print(f"\n{len(found)} findings, {blocking_count} blocking.")
    return 1 if blocking_count else 0


def _fixture(path, rows):
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(FIELDS)
    for row in rows:
        sheet.append([row.get(f, "") for f in FIELDS])
    book.save(path)


def selfcheck():
    """One runnable check, through the real file path: xlsx in, exit code out."""
    clean = dict.fromkeys(FIELDS, "None")
    clean.update({
        "Story ID": "MSP-001", "Type": "Story", "Epic": "Portal", "Phase": "Phase 1",
        "User Role": "Client User", "Priority (MoSCoW)": "Must", "Status": "Approved",
        "User Story": "As a Client User, I want to submit a request so that it is tracked",
        "Approved By": "MS Portal Owner, 6 May 2026",
        "Acceptance Criteria": "Given the ticketing system is unavailable, Then an error shows",
    })
    broken = dict(clean, **{
        "Story ID": "MSP-002", "Phase": "Phase 9", "Priority (MoSCoW)": "Won’t",
        "Dependencies": "MSP-404 and MSP-001", "Assumptions": "", "Approved By": "6 May 2026",
        "Acceptance Criteria": "Given a click, Then the page is fast",
    })
    stale = dict(clean, **{
        "Story ID": "MSP-003", "Status": "In Review",
        "Approved By": "Requested Portal Owner, 5 Jan 2026; proceeding at risk "
                       "per CH-001; revisit 1 Feb 2026",
    })
    joint = dict(clean, **{"Story ID": "MSP-004",
                           "Approved By": "Portal Owner 6 May 2026; Finance Lead pending"})
    quarantine = dict(clean, **{"Story ID": "MSP-005", "Status": "Draft",
                                "Assumptions": "unaudited, treat as unverified"})
    impossible = dict(clean, **{"Story ID": "MSP-007", "Status": "Draft",
                                "Assumptions": "unaudited, treat as unverified; "
                                               "revisit 30 Feb 2026"})
    depends_on_quarantined = dict(clean, **{"Story ID": "MSP-006", "Status": "Draft",
                                            "Dependencies": "MSP-005"})

    today = dt.date(2026, 8, 13)
    tmp = Path(tempfile.mkdtemp())

    _fixture(tmp / "bad.xlsx", [clean, broken, stale, joint, quarantine,
                                depends_on_quarantined, impossible])
    rows, headers, merged = load_rows(tmp / "bad.xlsx")
    assert len(rows) == 7, f"expected 7 rows through the file, got {len(rows)}"
    found = check(rows, headers, tier="2", merged=merged, today=today)
    blocking = [c for b, c, _ in found if b]
    causes = [c for _, c, _ in found]

    assert any("closed set" in c and "Phase 9" in c for c in blocking), "closed set not caught"
    assert not any("Won" in c for c in blocking), "curly apostrophe false positive"
    assert any("MSP-404 does not exist" in c for c in blocking), "dangling dependency not caught"
    assert not any("MSP-001 does not exist" in c for c in blocking), "valid dependency flagged"
    assert any("names no role" in c for c in blocking), "date without a role not caught"
    assert any("past its revisit date" in c for c in blocking), "stale debt record not caught"
    assert any("joint" in c for c in blocking), "half landed joint approval not caught"
    assert any("quarantined" in c for c in blocking), "quarantined dependency not caught"
    assert any("nothing brings it back" in c for c in blocking), \
        "a quarantine marker with no revisit date not caught"
    assert any("not a real date" in c for c in blocking), \
        "a date shaped but impossible revisit date not caught"
    assert any("blank Assumptions" in c for c in causes), "blank field not caught"
    assert any("'fast'" in c for c in causes), "vague word not caught"
    assert not any("heuristic" in c for c in blocking), "a heuristic was marked blocking"

    # The four structural blockers, none of which had an assertion before.
    structural = check([clean, dict(clean)], ["Story ID", "Story ID", "Notes"], merged=2)
    for expected in ("missing required columns", "duplicate column headers",
                     "merged cell range", "duplicate Story ID"):
        assert any(expected in c for b, c, _ in structural if b), \
            f"structural check gone: {expected}"
    # The fixtures are meant to fail; printing their report reads as a broken tool.
    with contextlib.redirect_stdout(io.StringIO()):
        assert report(found) == 1, "blocking findings must exit non zero"

    _fixture(tmp / "good.xlsx", [clean])
    rows, headers, merged = load_rows(tmp / "good.xlsx")
    good = check(rows, headers, merged=merged, today=today)
    assert not [c for b, c, _ in good if b], f"clean workbook produced blocking findings: {good}"
    with contextlib.redirect_stdout(io.StringIO()):
        assert report(good) == 0, "a clean workbook must exit zero"

    assert vague_words(), "the shared vague word list did not load"

    # Records written exactly as the method prescribes must not be findings.
    # Every one of these was a blocking false positive found by audit.
    prescribed = {
        "revisit used in prose": {"Description / Notes":
            "the Portal Owner asked us to revisit this wording after 1 Mar 2026"},
        "the yes lands, at risk record annotated": {"Approved By":
            "Portal Owner, 13 Aug 2026, covers work already built; superseded, "
            "see 04 CH-003"},
        "ba-change's supersede annotation": {"Approved By":
            "Portal Owner, 6 May 2026; superseded, see 04 CH-002"},
        "unaudited in its ordinary sense": {"Assumptions":
            "the 2026 figures are unaudited management accounts"},
        "a date written May 6, 2026": {"Approved By": "Portal Owner, May 6, 2026"},
        "revisit with no date near it": {"Description / Notes":
            "revisit once the vendor confirms, target sprint 6 2026"},
    }
    for label, patch in prescribed.items():
        blocking = [c for b, c, _ in check([dict(clean, **patch)], FIELDS, today=today) if b]
        assert not blocking, f"blocking false positive on {label}: {blocking}"

    # ...and the loosenings that bought those must not have opened a hole.
    must_block = {
        "an open at risk record on an Approved row": ({"Approved By":
            "Portal Owner, 6 May 2026; proceeding at risk per CH-009; revisit "
            "1 Dec 2026"}, "on an Approved row"),
        "a joint approver laundered with a CH handle": ({"Approved By":
            "Portal Owner 6 May 2026; Finance Lead pending, see 04 CH-001"},
            "carries no date"),
        "a hold whose revisit date has passed": ({"Status": "Deferred",
            "Description / Notes": "on hold pending the vendor, revisit 1 Feb 2026"},
            "past its revisit date"),
        "a quarantine marker punctuated differently": ({"Status": "Draft",
            "Assumptions": "unaudited - treat as unverified"},
            "nothing brings it back"),
    }
    for label, (patch, expected) in must_block.items():
        blocking = [c for b, c, _ in check([dict(clean, **patch)], FIELDS, today=today) if b]
        assert any(expected in c for c in blocking), f"{label} no longer blocks: {blocking}"

    # A deferred NFR is the accessibility target on an MVP clock: a debt record.
    # A held story is not, since it is parked by someone with the authority to park it.
    def deferred(prefix, **extra):
        return [dict(clean, **{"Story ID": f"{prefix}-{n}", "Status": "Deferred",
                               "Description / Notes": f"deferred per CH-10{n}, "
                                                      f"revisit 1 Dec 2026"}, **extra)
                for n in range(3)]

    assert any("open debt records" in c for _, c, _ in
               check(deferred("NFR", Type="NFR"), FIELDS, today=today)), \
        "three deferred NFRs are three debt records"
    assert not any("open debt records" in c for _, c, _ in
                   check(deferred("HLD"), FIELDS, today=today)), \
        "three held stories are not a debt pile"
    held = dict(clean, **{"Status": "Deferred",
                          "Description / Notes": "on hold per CH-001, revisit 1 Dec 2026"})
    assert any("quietly dying" in c for b, c, _ in
               check([dict(held, **{"Description / Notes": "on hold per CH-001"})],
                     FIELDS, today=today) if b), "a hold with no revisit date must block"
    # The word alone is not a debt record; the cited authority is what makes one.
    assert not [c for b, c, _ in check([dict(clean, **{"Approved By":
        "Portal Owner, 6 May 2026 - agreed to revisit at the pilot review"})],
        FIELDS, today=today) if b], "revisit as a verb in Approved By must not block"
    assert any("nothing brings it back" in c for b, c, _ in check([dict(clean, **{
        "Status": "Draft", "Approved By": "story owed per CH-008, revisit"})],
        FIELDS, today=today) if b), "a shipped fix debt with no date must block"

    # Formality is read from the Dials block, and Contract only reports.
    at_risk_row = dict(clean, **{"Status": "In Review", "Approved By":
        "Requested Portal Owner, 5 Aug 2026; no response as of 13 Aug 2026; "
        "proceeding at risk per CH-004; revisit 1 Dec 2026"})
    for setting in (None, "firm"):
        assert not [c for b, c, _ in
                    check([at_risk_row], FIELDS, today=today, formality=setting) if b], \
            f"an at risk record must not block on formality {setting}"
    contract = check([at_risk_row], FIELDS, today=today, formality="contract")
    assert any("Contract project" in c for _, c, _ in contract), "Contract said nothing"
    assert not [c for b, c, _ in contract if b], "Contract must report, not block"
    assert not [c for b, c, _ in
                check([clean], FIELDS, today=today, formality="contract") if b], \
        "an ordinary row on a Contract project must not block"
    kb = Path(__file__).resolve().parent.parent / "assets" / "project" / "knowledgebase"
    assert read_formality(kb) is None, "the template's bracketed formality must not parse"
    (tmp / "01_Requirements_Source.md").write_text(
        "## Dials\n\nSecurity tier: Tier 2\nFormality: Firm\n\n## Constraints\n\n"
        "A contract is being negotiated; formality: contract if it lands.\n",
        encoding="utf-8")
    assert read_formality(tmp) == "firm", "a dial must be read from the Dials block only"
    assert read_tier(tmp) == "2", "tier not read from the Dials block"

    # A tracker store exports to CSV and runs the identical sweep.
    # utf-8-sig on the way out too: a tracker export opened once in Excel comes
    # back with a BOM, and a BOM on cell one renames the Story ID column.
    with open(tmp / "export.csv", "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(FIELDS)
        for n in (7, 8, 9):
            row = dict(at_risk_row, **{"Story ID": f"MSP-{n:03d}",
                                       "Approved By": at_risk_row["Approved By"]
                                       .replace("CH-004", f"CH-{n:03d}")})
            writer.writerow([row.get(f, "") for f in FIELDS])
    # The same export as Excel writes it on a Western box: cp1252, curly
    # apostrophe. Mangling the byte turned a legal Won't into a blocking finding.
    (tmp / "cp1252.csv").write_bytes(
        (tmp / "export.csv").read_text(encoding="utf-8-sig")
        .replace("Must", "Won’t").encode("cp1252"))
    cp_rows, cp_headers, _ = load_rows(tmp / "cp1252.csv")
    assert cp_rows[0]["Priority (MoSCoW)"] == "Won’t", "cp1252 csv did not decode"
    assert not [c for b, c, _ in check(cp_rows, cp_headers, today=today) if b], \
        "a cp1252 export blocked on a legal value"

    rows, headers, merged = load_rows(tmp / "export.csv")
    assert len(rows) == 3, f"csv export did not read back as rows: {rows}"
    found = check(rows, headers, merged=merged, today=today)
    assert not [c for b, c, _ in found if b], f"future dated debt must not block: {found}"
    assert any("3 open debt records" in c for _, c, _ in found), "the pile was not counted"
    assert not any("open debt records" in c for _, c, _ in
                   check(rows[:2], headers, today=today)), "two records must not trip it"
    assert not any("open debt records" in c for _, c, _ in check(
        [dict(r, **{"Approved By": at_risk_row["Approved By"]}) for r in rows],
        headers, today=today)), "one directive across three rows is one record"

    matrix = [
        {"Requirement ID": "REQ-001", "Requirement": "The system shall log all access",
         "Response": "Comply", "How": "Native audit log", "Assumption": "None",
         "Reference": "S3.2"},
        {"Requirement ID": "REQ-002", "Requirement": "The portal must be intuitive",
         "Response": "Partial", "How": "Usability testing in scope", "Assumption": "None",
         "Reference": "None"},
        {"Requirement ID": "REQ-003", "Requirement": "Single sign on",
         "Response": "Yes", "How": "Supported", "Assumption": "None", "Reference": "None"},
    ]
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(MATRIX_FIELDS)
    for row in matrix:
        sheet.append([row.get(f, "") for f in MATRIX_FIELDS])
    book.save(tmp / "matrix.xlsx")
    rows, headers, _ = load_rows(tmp / "matrix.xlsx", key="Requirement ID")
    found = check_matrix(rows, headers)
    blocking = [c for b, c, _ in found if b]
    assert any("no assumption" in c for c in blocking), "Partial without an assumption not caught"
    assert any("closed set" in c for c in blocking), "invalid Response not caught"
    assert any("intuitive" in c for c, in [(c,) for _, c, _ in found]), \
        "vague word in the client's requirement not caught"
    assert not any("REQ-001" == rid for b, _, rid in found if b), "clean matrix row flagged"

    print("selfcheck passed")


def main():
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("workbook", nargs="?", help="path to the story workbook (.xlsx)")
    parser.add_argument("--kb", help="knowledgebase directory holding 01_Requirements_Source.md")
    parser.add_argument("--matrix", action="store_true",
                        help="validate a bid compliance matrix instead of a story workbook")
    parser.add_argument("--selfcheck", action="store_true", help="run the built in check and exit")
    args = parser.parse_args()

    if args.selfcheck:
        selfcheck()
        return 0
    if not args.workbook:
        parser.error("give a workbook path, or --selfcheck")
    if not Path(args.workbook).exists():
        parser.error(f"no workbook at {args.workbook}")
    if not vague_words():
        print(f"note: could not read the word list at {VAGUE_LIST}, "
              f"so the vague word scan is skipped\n")

    if args.matrix:
        if args.kb:
            print("note: --kb is not used when validating a matrix\n")
        rows, headers, merged = load_rows(args.workbook, key="Requirement ID")
        if not rows:
            print("no requirement rows found")
            return 1
        return report(check_matrix(rows, headers, merged=merged))

    tier = read_tier(args.kb) if args.kb else None
    formality = read_formality(args.kb) if args.kb else None
    if args.kb:
        print("dials: " + ", ".join(
            [f"security tier {tier}" if tier else "security tier unset",
             f"formality {formality}" if formality else "formality unset"]) + "\n")
    if args.kb and tier is None:
        print("note: no security tier found in the Dials block of "
              "01_Requirements_Source.md, so the tier 2 NFR check is skipped\n")

    rows, headers, merged = load_rows(args.workbook)
    if not rows:
        print("no story rows found")
        return 1
    return report(check(rows, headers, tier=tier, merged=merged, formality=formality))


if __name__ == "__main__":
    sys.exit(main())
