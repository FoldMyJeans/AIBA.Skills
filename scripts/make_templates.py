"""Generate the workbook templates in assets/.

    python scripts/make_templates.py
    python scripts/make_templates.py --verify

The generated files are committed, so a BA starting a project downloads and
opens them rather than running Python. Regenerate after changing a closed
vocabulary; the dropdowns are built from validate_workbook.py, so what a BA can
type and what the validator accepts come from one source. --verify checks that
the committed files still match it.
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent))
from validate_workbook import (CLOSED, FIELDS, MATRIX_FIELDS, MATRIX_RESPONSES,
                               check, check_matrix, load_rows)

ASSETS = Path(__file__).resolve().parent.parent / "assets"
STORY_TEMPLATE = ASSETS / "story_workbook_template.xlsx"
MATRIX_TEMPLATE = ASSETS / "compliance_matrix_template.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")

WIDTHS = {
    "Story ID": 12, "Date Created": 13, "Type": 11, "Epic": 18, "Phase": 13,
    "User Role": 16, "User Story": 52, "Description / Notes": 52,
    "Acceptance Criteria": 52, "Priority (MoSCoW)": 16, "Data Source": 18,
    "Dependencies": 16, "Assumptions": 30, "Open Questions": 30, "Status": 15,
    "Approved By": 30, "Estimated Hours": 15, "Estimate Notes": 34,
    "Requirement ID": 15, "Requirement": 60, "Response": 14, "How": 60,
    "Assumption": 40, "Reference": 16,
}


def style_header(sheet, headers):
    for index, name in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = WIDTHS.get(name, 20)
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"


def add_dropdown(sheet, headers, column_name, values, message):
    """A list validation, so a closed set stays closed at the point of typing."""
    if column_name not in headers:
        return
    letter = get_column_letter(headers.index(column_name) + 1)
    # Values keep their declared order: for Status that order is the lifecycle.
    formula = '"' + ",".join(values) + '"'
    if len(formula) > 255:
        raise ValueError(f"{column_name} list is too long for an inline Excel validation; "
                         f"move it to a lookup sheet")
    validation = DataValidation(type="list", formula1=formula, allow_blank=True,
                                showErrorMessage=True)
    validation.error = message
    validation.errorTitle = "Closed vocabulary"
    sheet.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}500")


def story_workbook(path):
    book = Workbook()
    sheet = book.active
    sheet.title = "Stories"
    headers = FIELDS + ["Estimated Hours", "Estimate Notes"]
    style_header(sheet, headers)
    for column, values in CLOSED.items():
        add_dropdown(sheet, headers, column, values,
                     f"{column} is a closed set. Pick from the list. Adding a value "
                     f"means changing the method, not just this project: declare it in "
                     f"02_Conventions.md and change CLOSED in validate_workbook.py.")
    sheet.append([
        "ABC-001", "2026-01-15", "Story", "Example Epic", "Phase 1", "Example Role",
        "As an Example Role, I want this row deleted so that the workbook starts clean",
        "The example row. It shows the format and passes the validator. Delete it "
        "before the first real story.",
        "Given the template is opened, When this row is deleted, Then the workbook is "
        "ready for the project. Given a required cell is left empty, Then the validator "
        "reports it as a finding.",
        "Must", "None", "None", "None", "None", "Draft", "", "N/A", "N/A",
    ])
    save(book, path)


def compliance_matrix(path):
    book = Workbook()
    sheet = book.active
    sheet.title = "Compliance Matrix"
    style_header(sheet, MATRIX_FIELDS)
    add_dropdown(sheet, MATRIX_FIELDS, "Response", MATRIX_RESPONSES,
                 "Response is a closed set. Pick from the list. Evaluators filter "
                 "this column, so a sentence here is not an answer.")
    sheet.append([
        "REQ-001", "The system shall record every access attempt.", "Comply",
        "Native audit log, retained 12 months. The example row: delete before use.",
        "None", "Proposal section 3.2",
    ])
    save(book, path)


def save(book, path):
    book.save(path)


def verify():
    """Do the committed templates still match the vocabularies they came from?

    An xlsx is a zip of timestamped members, so it never rehashes identically and
    a diff cannot answer this. What matters is narrower anyway: whether the
    dropdowns still offer exactly what the validator accepts.
    """
    import openpyxl

    problems = []
    expected = {
        STORY_TEMPLATE: (FIELDS + ["Estimated Hours", "Estimate Notes"],
                         {c: list(v) for c, v in CLOSED.items()}),
        MATRIX_TEMPLATE: (MATRIX_FIELDS, {"Response": list(MATRIX_RESPONSES)}),
    }
    for path, (headers, vocabularies) in expected.items():
        if not path.exists():
            problems.append(f"{path.name} is missing, so run this script")
            continue
        sheet = openpyxl.load_workbook(path).active
        actual = [c.value for c in sheet[1]]
        if actual != headers:
            problems.append(f"{path.name} headers differ from the field list: {actual}")
        # Keyed by column letter, because a dropdown offering the right values
        # on the wrong column is the regression a value-only check sails past.
        offered = {}
        for dv in sheet.data_validations.dataValidation:
            for cells in str(dv.sqref).split():
                offered[cells.split("2:")[0]] = (dv.formula1.strip('"'), str(dv.sqref))
        for column, values in vocabularies.items():
            if column not in headers:      # same guard add_dropdown carries
                continue
            letter = get_column_letter(headers.index(column) + 1)
            found, sqref = offered.get(letter, (None, ""))
            if found != ",".join(values):
                problems.append(f"{path.name} column {letter} ({column}) offers {found!r}, "
                                f"not {', '.join(values)}")
            elif sqref != f"{letter}2:{letter}500":
                problems.append(f"{path.name} {column} validation covers {sqref}, "
                                f"so some rows have no dropdown")

        # The example row is a claim the README makes; hold it to the validator.
        rows, actual_headers, merged = load_rows(path, key=headers[0])
        if not rows:
            problems.append(f"{path.name} has no example row, and the README says it has one")
        else:
            # Every finding, not just the blocking ones: the README says the row
            # passes, and "1 findings" is not passing.
            problems += [f"{path.name} example row does not pass the validator: {c}"
                         for _, c, _ in (check_matrix(rows, actual_headers, merged)
                                         if path == MATRIX_TEMPLATE
                                         else check(rows, actual_headers, merged=merged))]
    for problem in problems:
        print(problem)
    print("templates match the vocabularies" if not problems
          else f"\n{len(problems)} mismatch(es), regenerate with: python scripts/make_templates.py")
    return 1 if problems else 0


def main():
    if "--verify" in sys.argv:
        return verify()
    ASSETS.mkdir(exist_ok=True)
    story_workbook(STORY_TEMPLATE)
    compliance_matrix(MATRIX_TEMPLATE)
    print(f"wrote templates to {ASSETS}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
